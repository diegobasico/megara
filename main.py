import re
import duckdb
import polars as pl
from pathlib import Path
from openpyxl import load_workbook

# ----------------
# Paths
# ----------------

EXCEL_PATH = Path(r"C:\Users\diego\Desktop\AISC_profiles.xlsx")
DB_PATH = Path("aisc_sections.duckdb")

# ----------------
# Helpers
# ----------------


def clean_name(name: str) -> str:
    """
    Normalize Excel column names to snake_case.
    """
    name = name.strip().lower()
    name = re.sub(r"[^\w]+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def clean_table_name(name: str) -> str:
    name = name.strip().lower()
    name = re.sub(r"[^\w]+", "_", name)
    name = re.sub(r"_+", "_", name)
    return f"aisc_{name}"


# ----------------
# Sheet discovery (openpyxl)
# ----------------

wb = load_workbook(EXCEL_PATH, read_only=True)
sheet_names = wb.sheetnames
wb.close()

# ----------------
# Database
# ----------------

con = duckdb.connect(DB_PATH)

# ----------------
# Migration
# ----------------

for sheet in sheet_names:
    print(f"▶ Migrating sheet: {sheet}")

    df = pl.read_excel(EXCEL_PATH, sheet_name=sheet)

    if df.is_empty():
        print("  ↳ skipped (empty)")
        continue

    # Normalize columns
    df = df.rename({c: clean_name(c) for c in df.columns})
    df = df.filter(pl.any_horizontal(pl.all().is_not_null()))

    # Best-effort numeric casting
    df = df.with_columns([pl.col(c).cast(pl.Float64, strict=False) for c in df.columns])

    table_name = clean_table_name(sheet)

    con.execute(f"DROP TABLE IF EXISTS {table_name}")
    con.register("tmp_df", df)
    con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM tmp_df")
    con.unregister("tmp_df")

    print(f"  ↳ written → {table_name} ({df.height} rows)")

# ----------------
# Done
# ----------------

con.close()
print("✔ Migration completed successfully")
