import re
import logging
from pathlib import Path
from typing import Dict

import duckdb
import polars as pl
from openpyxl import load_workbook

# ----------------
# Logging
# ----------------

logger = logging.getLogger(__name__)


# ----------------
# Cleaning helpers
# ----------------


def regex_clean(name: str) -> str:
    """
    Normalize Excel names to snake_case.
    """
    name = name.strip().lower()
    name = re.sub(r"[^\w]+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def cast_inches(value: str | None) -> float | None:
    if value is None:
        return None

    value = value.strip()

    # Case: mixed fraction (e.g. 5-1/2)
    m = re.fullmatch(r"(\d+)-(\d+)/(\d+)", value)
    if m:
        whole, num, den = m.groups()
        return float(whole) + float(num) / float(den)

    # Case: decimal (e.g. 2.75)
    if re.fullmatch(r"\d+(\.\d+)?", value):
        return float(value)

    return None


# ----------------
# Sheet →  DataFrame
# ----------------


def load_and_clean_sheet(
    excel_path: Path,
    sheet_name: str,
) -> pl.DataFrame:
    df = pl.read_excel(excel_path, sheet_name=sheet_name)

    if df.is_empty():
        return df

    df = df.rename({c: regex_clean(c) for c in df.columns})

    # Drop rows where all values are null
    df = df.filter(pl.any_horizontal(pl.all().is_not_null()))

    # Trim strings
    df = df.with_columns(pl.col(pl.Utf8).str.strip_chars())

    return df


# ----------------
# Extract ALL sheets (no DB)
# ----------------


def extract_sheets(
    excel_path: Path,
) -> Dict[str, pl.DataFrame]:
    wb = load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    tables: Dict[str, pl.DataFrame] = {}

    for sheet in sheet_names:
        logger.info(f"▶ Loading sheet: {sheet}")

        df = load_and_clean_sheet(excel_path, sheet)

        if df.is_empty():
            logger.info("  ↳ skipped (empty)")
            continue

        table_name = f"aisc_{regex_clean(sheet)}"
        tables[table_name] = df

        logger.info(f"  ↳ loaded → {table_name} ({df.height} rows)")

    return tables


# ----------------
# Persistence (DB only)
# ----------------


def save_tables(
    tables: Dict[str, pl.DataFrame],
    db_path: Path,
    replace: bool = True,
):
    with duckdb.connect(db_path) as con:
        for table_name, df in tables.items():
            logger.info(f"▶ Writing table: {table_name}")

            if replace:
                con.execute(f"DROP TABLE IF EXISTS {table_name}")

            con.register("tmp_df", df)
            con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM tmp_df")
            con.unregister("tmp_df")

            logger.info(f"  ↳ written ({df.height} rows)")


def normalize_aisc_tables(
    tables: dict[str, pl.DataFrame],
) -> dict[str, pl.DataFrame]:
    # ----------------
    # Drop tables
    # ----------------
    drop_keys = {
        "aisc_aisc_properties_viewer_sim",
        "aisc_aisc_properties_viewer",
    }

    tables = {k: v for k, v in tables.items() if k not in drop_keys}

    # ----------------
    # Rename tables
    # ----------------
    rename_map = {
        "aisc_w_s_m_hp_shapes": "wsmhp",
        "aisc_c_mc_shapes": "cmc",
        "aisc_wt_shapes": "wt",
        "aisc_angles": "angles",
        "aisc_2angles": "two_angles",
        "aisc_tubes": "tubes",
        "aisc_pipes": "pipes",
    }

    out: dict[str, pl.DataFrame] = {}

    for old, df in tables.items():
        if old not in rename_map:
            raise KeyError(f"Unexpected table: {old}")

        out[rename_map[old]] = df
    return out


def normalize_wsmhp_table(df: pl.DataFrame) -> pl.DataFrame:
    df = df.with_columns(
        pl.col("gage").map_elements(cast_inches, return_dtype=pl.Float64).alias("gage"),
        pl.col("t").map_elements(cast_inches, return_dtype=pl.Float64).alias("t"),
        pl.col("fy").cast(pl.Float64, strict=False).alias("fy"),
    )

    return df


def normalize_cmc_table(df: pl.DataFrame) -> pl.DataFrame:
    df = df.with_columns(
        pl.col("gage").map_elements(cast_inches, return_dtype=pl.Float64).alias("gage"),
        pl.col("t").map_elements(cast_inches, return_dtype=pl.Float64).alias("t"),
    )

    return df


def normalize_angles_table(df: pl.DataFrame) -> pl.DataFrame:
    df = df.with_columns(
        pl.col("h").cast(pl.Float64, strict=False).alias("h"),
    )
    return df


# ----------------
# Entry point
# ----------------


def migrate_excel_to_duckdb(
    excel_path: Path,
    db_path: Path,
) -> Dict[str, pl.DataFrame]:
    tables = extract_sheets(excel_path)

    # --------------------------------
    # Hook: operate on tables HERE
    # --------------------------------
    # tables = normalize_units(tables)
    # tables = validate_schemas(tables)
    # ------------------------------

    tables = normalize_aisc_tables(tables)

    tables["wsmhp"] = normalize_wsmhp_table(tables["wsmhp"])
    tables["cmc"] = normalize_cmc_table(tables["cmc"])
    # wt tables was already normalized
    tables["angles"] = normalize_angles_table(tables["angles"])
    # two_angles was already normalized
    # tubes was already normalized
    # pipes was already normalized

    save_tables(tables, db_path)
    logger.info("✔ Migration completed successfully")

    return tables
